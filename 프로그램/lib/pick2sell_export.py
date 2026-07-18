"""픽투셀 엑셀 일괄등록 양식(data/pick2sell_template.xlsx)에 맞춰 결과를 채워넣는다."""

import os

import openpyxl

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "pick2sell_template.xlsx"
)
DATA_START_ROW = 4
MAX_ROWS = 500

COL_URL = 2
COL_TITLE = 3
COL_CATEGORY_COUPANG = 5
COL_CATEGORY_SS = 6
COL_EXTRA_MARGIN = 8
COL_TARGET_PRICE = 9
COL_MEMO = 10


def export_to_pick2sell_template(rows, output_path):
    """rows: [{"url":..., "title":..., "category_code_coupang":..., "category_code_ss":...,
    "extra_margin":..., "target_price":..., "memo":...}, ...]

    extra_margin(H컬럼, 배송비를 여기 넣음)과 target_price(I컬럼)는 둘 다 넣지 마세요 -
    픽투셀은 target_price가 있으면 extra_margin을 무시해버려요.
    """
    if len(rows) > MAX_ROWS:
        raise ValueError(f"픽투셀 양식은 한 파일에 최대 {MAX_ROWS}개까지만 담을 수 있어요. ({len(rows)}개 입력됨)")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["3.0"]

    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=COL_URL, value=row.get("url"))
        if row.get("title"):
            ws.cell(row=r, column=COL_TITLE, value=str(row["title"])[:100])
        if row.get("category_code_coupang"):
            ws.cell(row=r, column=COL_CATEGORY_COUPANG, value=str(row["category_code_coupang"]))
        if row.get("category_code_ss"):
            ws.cell(row=r, column=COL_CATEGORY_SS, value=str(row["category_code_ss"]))
        if row.get("extra_margin") is not None:
            ws.cell(row=r, column=COL_EXTRA_MARGIN, value=round(row["extra_margin"]))
        if row.get("target_price") is not None:
            ws.cell(row=r, column=COL_TARGET_PRICE, value=round(row["target_price"]))
        if row.get("memo"):
            ws.cell(row=r, column=COL_MEMO, value=str(row["memo"])[:1000])

    wb.save(output_path)
    return output_path
