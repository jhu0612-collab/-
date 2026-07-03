# -*- coding: utf-8 -*-
"""
슝 업로드용 - 구글시트 → 엑셀 자동 입력 툴 (최종 완성본)
- N열 "수강생"만 추출 (환불자 제외)
- 시트 목록: 구글시트 탭 순서대로 표시
- 새로고침 해도 선택했던 강좌 유지
- 클어 / 타이탄 리스트 선택 지원 (열 설정을 리스트별로 분리)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import csv
import io
import os
import sys
import threading
from openpyxl import load_workbook, Workbook
from datetime import datetime


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# ── 구글시트 설정 ──────────────────────────────────────────
SPREADSHEET_ID = "1qclrbo3_VG-sSNIqMW4j1juzwP3nq_ZaT-y1z6WLafc"

# ── 리스트별 열 설정 ────────────────────────────────────────
# 클어/타이탄이 서로 다른 열 배치를 쓰게 되면 여기 숫자만 바꾸면 됩니다.
# (0-based 컬럼 인덱스: A=0, B=1, C=2 ...)
LIST_CONFIGS = {
    "클어": {
        "col_name": 2,     # C열 (이름/수강생)
        "col_m": 12,       # M열 (카톡방입장, 공백이어야 추출)
        "col_n": 13,       # N열 (수강생 여부, n_target 값이어야 추출)
        "col_p": 15,       # P열 (반배정)
        "n_target": "수강생",
    },
    "타이탄": {
        "col_name": 2,     # C열 (수강생)
        "col_m": 12,       # M열 (카톡방 입장)
        "col_n": 13,       # N열 (수강생)
        "col_p": 15,       # P열 (반배정)
        "n_target": "수강생",
    },
}
LIST_NAMES = list(LIST_CONFIGS.keys())

# ── 색상 테마 ──────────────────────────────────────────────
BG_MAIN = "#F0F4F8"
BG_CARD = "#FFFFFF"
BG_HEADER = "#4A6FA5"
FG_HEADER = "#FFFFFF"
BG_BUTTON = "#4A6FA5"
BG_BUTTON_HOVER = "#3A5A8C"
FG_BUTTON = "#FFFFFF"
BG_REFRESH = "#E8EEF4"
ACCENT = "#4A6FA5"
BORDER_COLOR = "#D1DBE6"
FONT_TITLE = ("맑은 고딕", 18, "bold")
FONT_SUBTITLE = ("맑은 고딕", 10)
FONT_LABEL = ("맑은 고딕", 10, "bold")
FONT_ENTRY = ("맑은 고딕", 11)
FONT_STATUS = ("맑은 고딕", 10)
FONT_SECTION = ("맑은 고딕", 11, "bold")


def get_sheet_names():
    """구글시트의 실제 탭 순서대로 시트 이름을 가져옵니다."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}"
        f"/gviz/tq?tqx=out:csv&sheet=__탭인덱스"
    )
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        reader = csv.reader(io.StringIO(resp.text))
        rows = list(reader)
        names = []
        for row in rows[1:]:
            if row and row[0].strip():
                names.append(row[0].strip())
        return names
    except Exception as e:
        messagebox.showerror("오류", f"시트 목록을 가져올 수 없습니다.\n{e}")
        return []


def get_sheet_names_from_tabs():
    """구글시트에서 실제 탭 순서대로 시트 이름을 가져옵니다."""
    import re
    url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()

        from openpyxl import load_workbook
        import tempfile
        tmp = os.path.join(tempfile.gettempdir(), "_temp_sheets.xlsx")
        with open(tmp, "wb") as f:
            f.write(resp.content)

        wb = load_workbook(tmp, read_only=True)
        xlsx_names = wb.sheetnames
        wb.close()
        os.remove(tmp)

        # __탭인덱스에서 원본 이름 목록 가져오기
        original_names = get_sheet_names()

        exclude = {"메인", "__탭인덱스", "C클래스 견본시트", "T클래스 견본시트"}
        xlsx_filtered = [n for n in xlsx_names if n not in exclude]

        if not xlsx_filtered or not original_names:
            return original_names if original_names else xlsx_filtered

        # xlsx 이름 → 원본 이름 매칭 (대괄호 제거된 이름으로 비교)
        def normalize(name):
            return re.sub(r'[\[\]/]', '', name).replace('  ', ' ').strip()

        orig_map = {}
        for orig in original_names:
            orig_map[normalize(orig)] = orig

        result = []
        seen = set()
        for xlsx_name in xlsx_filtered:
            key = normalize(xlsx_name)
            if key in orig_map and orig_map[key] not in seen:
                result.append(orig_map[key])
                seen.add(orig_map[key])

        # 매칭 안 된 원본 이름도 뒤에 추가
        for orig in original_names:
            if orig not in seen and orig not in exclude:
                result.append(orig)
                seen.add(orig)

        if result:
            return result

        return original_names
    except Exception:
        return get_sheet_names()



def fetch_data(sheet_name, list_name):
    config = LIST_CONFIGS[list_name]
    col_name = config["col_name"]
    col_m = config["col_m"]
    col_n = config["col_n"]
    col_p = config["col_p"]
    n_target = config["n_target"]

    url = (
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}"
        f"/gviz/tq?tqx=out:csv&sheet={requests.utils.quote(sheet_name)}"
    )
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    resp.encoding = "utf-8"
    reader = csv.reader(io.StringIO(resp.text))
    rows = list(reader)

    if not rows:
        return {}

    header = [h.strip() for h in rows[0]]

    col_phone = None
    for i, h in enumerate(header):
        h_clean = h.replace(" ", "")
        if "전화번호" in h_clean and col_phone is None:
            col_phone = i

    if col_phone is None:
        raise ValueError(f"시트에서 '전화번호' 열을 찾을 수 없습니다.")

    results_by_class = {}
    max_col_needed = max(col_name, col_phone, col_m, col_n, col_p)

    for row in rows[1:]:
        if len(row) <= max_col_needed:
            row += [''] * (max_col_needed - len(row) + 1)

        # M열: 공백인 사람만 추출 (뭐라도 적혀있으면 스킵)
        is_entered = row[col_m].strip()
        if is_entered != "":
            continue

        # N열: n_target 값만 추출 (환불자 등 제외)
        n_val = row[col_n].strip()
        if n_val != n_target:
            continue

        name = row[col_name].strip()
        phone = row[col_phone].strip()
        class_name = row[col_p].strip()

        if name and phone:
            if not class_name:
                class_name = "미배정"

            if class_name not in results_by_class:
                results_by_class[class_name] = []

            results_by_class[class_name].append({"name": name, "phone": phone})

    return results_by_class


def save_to_excel(data, course_name, entry_code, link_name, save_path):
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
    else:
        wb = Workbook()

    target_sheet = "발송 데이터"
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
    else:
        ws = wb.active
        ws.title = target_sheet

    ws.delete_rows(1, ws.max_row + 1)

    headers = ["연락처", "#{고객명}", "#{강좌명}", "#{입장코드}", "#{링크명}", "", ""]
    ws.append(headers)

    for item in data:
        ws.append([
            item["phone"],
            item["name"],
            course_name,
            entry_code,
            link_name,
            "",
            "",
        ])

    wb.save(save_path)
    return len(data)


class CardFrame(tk.Frame):
    def __init__(self, parent, title="", **kwargs):
        super().__init__(parent, bg=BG_MAIN, **kwargs)
        self.card = tk.Frame(self, bg=BG_CARD, highlightbackground=BORDER_COLOR,
                             highlightthickness=1, padx=18, pady=14)
        self.card.pack(fill="x", padx=2, pady=2)

        if title:
            title_label = tk.Label(self.card, text=title, font=FONT_SECTION,
                                   bg=BG_CARD, fg=ACCENT, anchor="w")
            title_label.pack(fill="x", pady=(0, 8))
            separator = tk.Frame(self.card, bg=BORDER_COLOR, height=1)
            separator.pack(fill="x", pady=(0, 10))

    def get_inner(self):
        return self.card


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("슝 업로드용 - 구글시트 → 엑셀 자동 입력")
        self.root.geometry("820x780")
        self.root.resizable(False, False)
        self.root.configure(bg=BG_MAIN)

        header_frame = tk.Frame(root, bg=BG_HEADER, height=80)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        header_inner = tk.Frame(header_frame, bg=BG_HEADER)
        header_inner.place(relx=0.5, rely=0.5, anchor="center")

        text_and_img_frame = tk.Frame(header_inner, bg=BG_HEADER)
        text_and_img_frame.pack()

        text_frame = tk.Frame(text_and_img_frame, bg=BG_HEADER)
        text_frame.pack(side="left", padx=(0, 20))

        tk.Label(text_frame, text="슝 업로드용 자동화 툴", font=FONT_TITLE, bg=BG_HEADER, fg=FG_HEADER).pack()
        tk.Label(text_frame, text="M열(입장필터) + N열(수강생필터) + P열(반별 분할 저장) 지원", font=FONT_SUBTITLE, bg=BG_HEADER, fg="#C8D8E8").pack(pady=(2, 0))

        try:
            from PIL import Image, ImageTk
            img_path = resource_path("dog.png")
            dog_image = Image.open(img_path)
            dog_image = dog_image.resize((60, 60), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(dog_image)

            dog_label = tk.Label(text_and_img_frame, image=photo, bg=BG_HEADER)
            dog_label.image = photo
            dog_label.pack(side="left")
        except Exception as e:
            tk.Label(text_and_img_frame, text="(강아지 사진 없음)", bg=BG_HEADER, fg="#C8D8E8").pack(side="left")

        main = tk.Frame(root, bg=BG_MAIN, padx=20, pady=10)
        main.pack(fill="both", expand=True)

        card0 = CardFrame(main, title="① 리스트 선택")
        card0.pack(fill="x", pady=(0, 10))
        row0 = tk.Frame(card0.get_inner(), bg=BG_CARD)
        row0.pack(fill="x")

        self.list_var = tk.StringVar(value=LIST_NAMES[0])
        for name in LIST_NAMES:
            tk.Radiobutton(
                row0, text=name, variable=self.list_var, value=name,
                font=FONT_ENTRY, bg=BG_CARD, activebackground=BG_CARD,
                selectcolor=BG_REFRESH, cursor="hand2",
            ).pack(side="left", padx=(0, 20))
        # 클어/타이탄은 같은 구글시트를 쓰므로, 리스트를 바꿔도 시트 목록을
        # 다시 받아올 필요가 없다. (열 설정 차이는 실행 시점에만 적용됨)

        card1 = CardFrame(main, title="② 시트(강좌) 선택")
        card1.pack(fill="x", pady=(0, 10))
        row1 = tk.Frame(card1.get_inner(), bg=BG_CARD)
        row1.pack(fill="x")

        self.combo_sheet = ttk.Combobox(row1, state="readonly", width=40, font=FONT_ENTRY)
        self.combo_sheet.pack(side="left", padx=(0, 10))

        self.btn_refresh = tk.Button(row1, text="⟳ 새로고침", font=("맑은 고딕", 10), bg=BG_REFRESH, fg="#4A6FA5", relief="flat",
                                cursor="hand2", padx=12, pady=4, command=self.refresh_sheets)
        self.btn_refresh.pack(side="left")

        card2 = CardFrame(main, title="③ 반별 입력 항목 (P열 '반배정' 이름과 똑같이 적어주세요!)")
        card2.pack(fill="x", pady=(0, 10))
        inner2 = card2.get_inner()

        grid_frame = tk.Frame(inner2, bg=BG_CARD)
        grid_frame.pack(fill="x")

        headers = ["대상 반 이름(P열)", "강좌명", "입장코드", "링크명"]
        for col, text in enumerate(headers):
            tk.Label(grid_frame, text=text, font=FONT_LABEL, bg=BG_CARD, fg="#333").grid(row=0, column=col, padx=5, pady=5)

        self.class_entries = []
        for i in range(5):
            e_cls = tk.Entry(grid_frame, width=15, font=FONT_ENTRY, relief="solid", bd=1)
            e_cls.grid(row=i + 1, column=0, padx=5, pady=4, ipady=3)

            e_course = tk.Entry(grid_frame, width=25, font=FONT_ENTRY, relief="solid", bd=1)
            e_course.grid(row=i + 1, column=1, padx=5, pady=4, ipady=3)

            e_code = tk.Entry(grid_frame, width=15, font=FONT_ENTRY, relief="solid", bd=1)
            e_code.grid(row=i + 1, column=2, padx=5, pady=4, ipady=3)

            e_link = tk.Entry(grid_frame, width=25, font=FONT_ENTRY, relief="solid", bd=1)
            e_link.grid(row=i + 1, column=3, padx=5, pady=4, ipady=3)

            self.class_entries.append({"cls": e_cls, "course": e_course, "code": e_code, "link": e_link})

        card3 = CardFrame(main, title="④ 엑셀 파일들이 저장될 폴더 선택")
        card3.pack(fill="x", pady=(0, 10))
        inner3 = card3.get_inner()
        row3 = tk.Frame(inner3, bg=BG_CARD)
        row3.pack(fill="x")

        self.save_dir_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop"))
        entry_path = tk.Entry(row3, textvariable=self.save_dir_var, width=50, font=("맑은 고딕", 10), relief="solid", bd=1)
        entry_path.pack(side="left", padx=(0, 10), ipady=4)

        btn_browse = tk.Button(row3, text="📁 폴더 찾기", font=("맑은 고딕", 10), bg=BG_REFRESH, fg="#4A6FA5", relief="flat",
                               cursor="hand2", padx=12, pady=4, command=self.browse_dir)
        btn_browse.pack(side="left")

        self.btn_run = tk.Button(main, text="▶  실행하기 (개별 엑셀 생성)", font=("맑은 고딕", 14, "bold"),
                                 bg=BG_BUTTON, fg=FG_BUTTON, relief="flat", cursor="hand2", pady=15,
                                 command=self.run)
        self.btn_run.pack(fill="x", pady=(10, 0))

        self.status_var = tk.StringVar(value="시트 목록을 불러오는 중...")
        self.status_label = tk.Label(main, textvariable=self.status_var, font=FONT_STATUS, bg=BG_MAIN, fg="#888888")
        self.status_label.pack(pady=5)

        self.root.after(500, self.refresh_sheets)

    def refresh_sheets(self):
        # 현재 선택된 강좌 이름 기억
        current_selection = self.combo_sheet.get()

        self.status_var.set("시트 목록을 불러오는 중...")
        self.status_label.config(fg="#F57C00")
        self.btn_refresh.config(state="disabled")

        def worker():
            names = get_sheet_names_from_tabs()
            self.root.after(0, lambda: self._apply_sheet_names(names, current_selection))

        threading.Thread(target=worker, daemon=True).start()

    def _apply_sheet_names(self, names, current_selection):
        self.btn_refresh.config(state="normal")
        if names:
            self.combo_sheet["values"] = names

            # 이전에 선택했던 강좌가 있으면 그대로 유지
            if current_selection and current_selection in names:
                idx = names.index(current_selection)
                self.combo_sheet.current(idx)
            else:
                self.combo_sheet.current(0)

            self.status_var.set(f"✔ 시트 {len(names)}개 로드 완료")
            self.status_label.config(fg="#2E7D32")
        else:
            self.status_var.set("✖ 시트 목록 로드 실패")
            self.status_label.config(fg="#C62828")

    def browse_dir(self):
        path = filedialog.askdirectory(title="엑셀 파일들을 저장할 폴더 선택")
        if path:
            self.save_dir_var.set(path)

    def run(self):
        list_name = self.list_var.get()
        sheet = self.combo_sheet.get()
        save_dir = self.save_dir_var.get().strip()

        if not sheet:
            messagebox.showwarning("입력 확인", "시트를 선택해주세요.")
            return
        if not save_dir:
            messagebox.showwarning("입력 확인", "저장할 폴더를 선택해주세요.")
            return

        active_configs = []
        for config in self.class_entries:
            cls_name = config["cls"].get().strip()
            if cls_name:
                active_configs.append({
                    "cls": cls_name,
                    "course": config["course"].get().strip(),
                    "code": config["code"].get().strip(),
                    "link": config["link"].get().strip()
                })

        if not active_configs:
            messagebox.showwarning("입력 확인", "최소 1개 이상의 대상 반 정보를 입력해주세요.")
            return

        self.status_var.set(f"[{list_name}] '{sheet}' 시트에서 데이터를 가져오는 중...")
        self.status_label.config(fg="#F57C00")
        self.root.update()

        try:
            data_by_class = fetch_data(sheet, list_name)

            if not data_by_class:
                messagebox.showinfo("결과", "조건에 맞는 수강생 데이터가 없습니다.\n(M열 공백 + N열 '수강생'인 사람만 추출)")
                self.status_var.set("데이터 없음")
                self.status_label.config(fg="#888888")
                return

            total_saved_count = 0
            created_files_msg = []

            for config in active_configs:
                target_class = config["cls"]

                if target_class in data_by_class:
                    target_data = data_by_class[target_class]
                    file_name = f"슝업로드용_{target_class}.xlsx"
                    file_path = os.path.join(save_dir, file_name)

                    count = save_to_excel(
                        data=target_data,
                        course_name=config["course"],
                        entry_code=config["code"],
                        link_name=config["link"],
                        save_path=file_path
                    )

                    total_saved_count += count
                    created_files_msg.append(f" - {target_class}: {count}명 저장 완료")
                else:
                    created_files_msg.append(f" - {target_class}: 데이터 없음 (스킵)")

            msg = "\n".join(created_files_msg)
            self.status_var.set(f"✔ 완료! 총 {total_saved_count}명 처리됨")
            self.status_label.config(fg="#2E7D32")
            messagebox.showinfo("처리 완료", f"[{list_name}] 총 {total_saved_count}명의 데이터를 엑셀로 분할 저장했습니다!\n\n[처리 내역]\n{msg}")

        except Exception as e:
            self.status_var.set("✖ 오류 발생")
            self.status_label.config(fg="#C62828")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다.\n\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
